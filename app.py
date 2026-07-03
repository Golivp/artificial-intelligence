import os
import json
import uuid
import time
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename

from config import UPLOAD_FOLDER, RESULTS_FOLDER, MODELS, DEFAULT_MODEL, DEFAULT_CONFIDENCE, DEFAULT_IOU
from database import init_db, insert_analysis, get_all_analyses, get_analysis_by_id
from yolo_utils import get_available_models, process_image, process_video, allowed_file

# Импорты для отчётов
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB
CORS(app)

# Инициализация БД
init_db()

# Глобальное состояние активной модели
current_model = DEFAULT_MODEL

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_models', methods=['GET'])
def get_models():
    """Возвращает список доступных моделей с их описаниями"""
    models_info = {name: desc for name, desc in MODELS.items()}
    return jsonify({
        'models': models_info,
        'current': current_model
    })

@app.route('/set_model', methods=['POST'])
def set_model():
    """Переключает активную модель"""
    global current_model
    data = request.get_json()
    model_name = data.get('model')
    if model_name not in MODELS:
        return jsonify({'error': 'Недопустимая модель'}), 400
    current_model = model_name
    return jsonify({'status': 'ok', 'current_model': current_model})

@app.route('/process', methods=['POST'])
def process_file():
    """Обработка загруженного файла (изображение или видео)"""
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не передан'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Неподдерживаемый формат файла'}), 400

    # Параметры из запроса
    model_name = request.form.get('model', current_model)
    confidence = float(request.form.get('confidence', DEFAULT_CONFIDENCE))
    iou = float(request.form.get('iou', DEFAULT_IOU))

    # Сохраняем загруженный файл
    original_filename = secure_filename(file.filename)
    ext = original_filename.rsplit('.', 1)[-1].lower()
    unique_filename = f"{uuid.uuid4().hex}_{int(time.time())}.{ext}"
    file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
    file.save(file_path)

    # Определяем тип файла
    if ext in {'png', 'jpg', 'jpeg', 'bmp', 'gif'}:
        file_type = 'image'
        result = process_image(file_path, model_name, confidence, iou)
    elif ext in {'mp4', 'avi', 'mov', 'mkv'}:
        file_type = 'video'
        result = process_video(file_path, model_name, confidence, iou)
    else:
        return jsonify({'error': 'Неподдерживаемый тип файла'}), 400

    # Сохраняем запись в БД
    analysis_id = insert_analysis(
        file_type=file_type,
        people_count=result['people_count'],
        file_path=file_path,
        processing_time=result['processing_time'],
        model_used=model_name,
        confidence=confidence,
        iou=iou,
        result_image_path=result.get('result_image_path'),
        video_stats=json.dumps(result.get('video_stats')) if result.get('video_stats') else None
    )

    # Формируем ответ
    response = {
        'analysis_id': analysis_id,
        'people_count': result['people_count'],
        'processing_time': result['processing_time'],
        'file_type': file_type,
        'model_used': model_name,
        'confidence': confidence,
        'iou': iou
    }
    if file_type == 'image' and result.get('result_image_path'):
        response['result_image_url'] = f'/results/{result["result_image_path"]}'
    if file_type == 'video' and result.get('video_stats'):
        response['video_stats'] = result['video_stats']

    return jsonify(response)

@app.route('/results/<filename>')
def get_result_image(filename):
    """Отдаёт обработанное изображение"""
    return send_from_directory(RESULTS_FOLDER, filename)

@app.route('/history', methods=['GET'])
def history():
    """Возвращает историю анализов"""
    analyses = get_all_analyses(limit=50)
    for a in analyses:
        if a.get('video_stats'):
            a['video_stats'] = json.loads(a['video_stats'])
    return jsonify(analyses)

@app.route('/report/<format>/<int:analysis_id>', methods=['GET'])
def generate_report(format, analysis_id):
    """Генерация отчёта в PDF или Excel"""
    analysis = get_analysis_by_id(analysis_id)
    if not analysis:
        return jsonify({'error': 'Анализ не найден'}), 404

    if format == 'pdf':
        return generate_pdf(analysis)
    elif format == 'excel':
        return generate_excel(analysis)
    else:
        return jsonify({'error': 'Неподдерживаемый формат'}), 400

def generate_pdf(analysis):
    """Создание PDF-отчёта с помощью ReportLab"""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Заголовок
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, height - 50, "Отчёт по анализу гостей за столом")
    c.setFont("Helvetica", 12)
    c.drawString(100, height - 80, f"ID анализа: {analysis['id']}")
    c.drawString(100, height - 100, f"Дата: {analysis['timestamp']}")
    c.drawString(100, height - 120, f"Тип файла: {analysis['file_type']}")
    c.drawString(100, height - 140, f"Количество людей: {analysis['people_count']}")
    c.drawString(100, height - 160, f"Модель: {analysis['model_used']}")
    c.drawString(100, height - 180, f"Уверенность: {analysis['confidence']}")
    c.drawString(100, height - 200, f"IOU: {analysis['iou']}")
    c.drawString(100, height - 220, f"Время обработки: {analysis['processing_time']:.2f} сек")

    # Если есть статистика видео
    if analysis.get('video_stats'):
        stats = json.loads(analysis['video_stats'])
        y = 240
        c.drawString(100, height - y, f"Среднее кол-во людей: {stats['avg_people']}")
        y += 20
        c.drawString(100, height - y, f"Максимум: {stats['max_people']}")
        y += 20
        c.drawString(100, height - y, f"Минимум: {stats['min_people']}")
        y += 20
        c.drawString(100, height - y, f"Анализировано кадров: {stats['analyzed_frames']} из {stats['total_frames']}")

    # Если есть изображение результата
    if analysis.get('result_image_path'):
        img_path = os.path.join(RESULTS_FOLDER, analysis['result_image_path'])
        if os.path.exists(img_path):
            try:
                img = ImageReader(img_path)
                # Масштабируем изображение
                img_width, img_height = img.getSize()
                max_width = 400
                max_height = 400
                scale = min(max_width/img_width, max_height/img_height, 1.0)
                draw_width = img_width * scale
                draw_height = img_height * scale
                c.drawImage(img, (width - draw_width)/2, height - 500 - draw_height,
                            width=draw_width, height=draw_height, preserveAspectRatio=True)
            except:
                pass

    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"report_{analysis['id']}.pdf",
                     mimetype='application/pdf')

def generate_excel(analysis):
    """Создание Excel-отчёта с помощью OpenPyXL"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовки
    headers = ["Параметр", "Значение"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    data = [
        ["ID анализа", analysis['id']],
        ["Дата", analysis['timestamp']],
        ["Тип файла", analysis['file_type']],
        ["Количество людей", analysis['people_count']],
        ["Модель", analysis['model_used']],
        ["Уверенность", analysis['confidence']],
        ["IOU", analysis['iou']],
        ["Время обработки (сек)", f"{analysis['processing_time']:.2f}"]
    ]

    if analysis.get('video_stats'):
        stats = json.loads(analysis['video_stats'])
        data.append(["Среднее кол-во людей", stats['avg_people']])
        data.append(["Максимум", stats['max_people']])
        data.append(["Минимум", stats['min_people']])
        data.append(["Анализировано кадров", f"{stats['analyzed_frames']} из {stats['total_frames']}"])

    row = 2
    for item in data:
        ws.cell(row=row, column=1, value=item[0])
        ws.cell(row=row, column=2, value=item[1])
        row += 1

    # Автоширина колонок
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"report_{analysis['id']}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Ресурс не найден'}), 404

@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': 'Внутренняя ошибка сервера'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
